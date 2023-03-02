from PyMovieDb import IMDB
import json
import easygui
from operator import itemgetter
from openpyxl import load_workbook
wb = load_workbook(filename = "Movies.xlsx")
base = wb["Watched"]
metadata = wb["MetaData"]
rowCount = 2
for row in range(2,base.max_row+1):  
    done = False
    while not done:
        sheet_title = "{}{}".format('C', row)
        sheet_year = "{}{}".format('E', row)
        if not base[sheet_title].value:
            break
        else:  
            try:
                name = metadata.cell(row = rowCount, column = 1)
                rating = metadata.cell(row = rowCount, column = 2)
                contentRating = metadata.cell(row = rowCount, column = 3)
                releaseDate = metadata.cell(row = rowCount, column = 4)
                genre = metadata.cell(row = rowCount, column = 5)
                directors = metadata.cell(row = rowCount, column = 7)
                actors = metadata.cell(row = rowCount, column = 6)
                creators = metadata.cell(row = rowCount, column = 8)
                if rating.value != 'data not found' and rating.value is not None:
                    rowCount = rowCount + 1
                    break
                try:
                    imdb = IMDB()
                    title = str(base[sheet_title].value)
                    year = int(base[sheet_year].value)
                    searched = imdb.search(title, year, False, False)
                    result = json.loads(searched)
                    print(title)
                    ourmovie = result['results'][0]
                except IndexError:
                    print("Title not found")
                movieString = imdb.get_by_id(ourmovie['id'])
                movieObject = json.loads(movieString)
                print(movieObject)
                name.value = movieObject['name']
                rating.value = movieObject['rating']['ratingValue']
                contentRating.value = movieObject['contentRating']
                releaseDate.value = movieObject['datePublished']
                genre.value = ','.join(movieObject['genre'])
                actors.value = ','.join(list(map(itemgetter('name'), movieObject['actor'])))
                directors.value = ','.join(list(map(itemgetter('name'), movieObject['director'])))
                creators.value = ','.join(list(map(itemgetter('name'), movieObject['creator'])))
                wb.save(filename = "Movies.xlsx")
                done = True
                print('Finished')
                rowCount = rowCount + 1
            except KeyError as e:
                print ('I got a KeyError - reason "%s"' % str(movieObject))
                name = metadata.cell(row = rowCount, column = 1)
                name.value = title
                rating = metadata.cell(row = rowCount, column = 2)
                rating.value = "data not found"
                wb.save(filename = "Movies.xlsx")
                rowCount = rowCount + 1
                done = True
input('Press ENTER to exit')


            #print(ourmovie.data)




            #row = metadata.get_highest_row() + 1
            #new_row = [ourmovie['title'], 'data2', 'data3', 'data4']
            #for col, entry in enumerate(new_row, start=1):
            #    ws.cell(row=row, column=col, value=entry)
